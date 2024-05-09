from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor
import time
import psutil
import multiprocessing
import openseespy.opensees as op
import os
import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from scipy.stats import lognorm, norm, uniform
import math
from datetime import datetime
import vfo.vfo as vfo
import BraineryWiz as BSV
global Epsfactor; Epsfactor=10
global factorAngle; factorAngle=1
global AttackAngle; AttackAngle=0


import opsvis as vis
global DOFsToMatch;DOFsToMatch=[1,2,3,4,5,6]
global DOFsToMatchc; DOFsToMatchc=[1,2,3]

#***********************************************************************************************************************
ram_info = psutil.virtual_memory()
print(f"System Total RAM: {ram_info.total / (1024 ** 3):.2f} GB")
cores= multiprocessing.cpu_count()
print(f"System Total Number of CPU cores: {cores}")
practiceCores= cores-2
print(f"Taking {practiceCores} Cores for Run!")
#***********************************************************************************************************************
#Model Criteria
global iTower; iTower=1
global cable; cable="on"
# Domain
global Nodeiloc, C_Elementiloc,B_Elementiloc

Nodeiloc=["Leg"] #,'Body','Window','CrossArm']
C_Elementiloc=['LegColumn']#,'LegBody','LegWindow']
B_Elementiloc=['LegBeam'] #,'Pylon','BeamBody','Neck','BeamWindow', 'Cage']

# global cableR; global cableL;global cableM  
# cableR=[[300001,  300002,  300003,  300004,  300005,  300006,  300007,  300008,  300009,  300010,  300011,  300012, 300013, 300014,  300015,  300016,  300017,  300018,  300019],
#         [600002,  600004,  600006,  600008,  600010,  600012,  600014,  600016,  600018,  600020,  600022, 600024,  600026,  600028,  600030,  600032,  600034,  600036,  600038]]
# cableL=[[300201,  300202,  300203,  300204,  300205,  300206,  300207,  300208,  300209,  300210,  300211,  300212,  300213,  300214,  300215,  300216,  300217,  300218,  300219],
#         [600402,  600404,  600406,  600408,  600410,  600412,  600414,  600416,  600418,  600420,  600422,  600424,  600426,  600428, 600430,  600432,  600434,  600436,  600438]]
# cableM=[[300401,  300402,  300403,  300404,  300405,  300406,  300407,  300408,  300409,  300410,  300411,  300412,  300413,  300414, 300415,  300416,  300417,  300418,  300419],
#         [600802,  600804,  600806,  600808,  600810,  600812,  600814,  600816,  600818,  600820,  600822,  600824,  600826, 600828,  600830,  600832,  600834,  600836,  600838]]
#***********************************************************************************************************************
global Attempt; Attempt=f"out_{datetime.now().date()}"
if cable=="on":
    Attempt=f"{Attempt}_With Cable"    
#Dynamic parameters
global timestep; timestep=0.1
global duration; duration=3600
global Record_deltaT; Record_deltaT=0.1
#Run Condition
global Reliability; Reliability='on'
global Output; Output='off'
global runplot; runplot='off'
global testplot; testplot='off'
global modeShape; modeShape='off'
global Runmodal; Runmodal='off'
global plotwindspeedprofile; plotwindspeedprofile='off'
global plotsection; plotsection='off'
#***********************************************************************************************************************
global SectionsInfos;
with open('Input/SectionInfos.txt') as file: 
    SectionsInfos=[]
    SectionsInfos=[list(map(float, line.split())) for line in file]
if Reliability=='off': 
    if not os.path.exists(f"Out_{Attempt}"):
        os.mkdir(f"Out_{Attempt}")
    op.logFile(f'Out_{Attempt}/log.txt')
    
    # if not os.path.exists(f"OutPush_{Attempt}"):
    #     os.mkdir(f"OutPush_{Attempt}")
    if not os.path.exists(f"Outsteps_{Attempt}"):
        os.mkdir(f"Outsteps_{Attempt}")
#***********************************************************************************************************************
#Model Function Input
global nMode; nMode=15
global N_Bmesh;N_Bmesh=3
global N_Cmesh; N_Cmesh=4
global NintegrationB; NintegrationB=3 
global NintegrationC; NintegrationC= 3
global CableN_seg; CableN_seg=20 # Number of Segments of the cable even number is better
global NumNodeDomain; NumNodeDomain=10000
#Cable and Insulator info
global g; g=9.806

#***********************************************************************************************************************
#Functions
def runWindsteps():
    dispx=[];dispy=[];reactx=[];reacty=[];wind=[]
    for i in range(20,106,5):
        WindSpeed=i/10

        CelementR, CelementL, CelementM, cableR, cableM,cableL,InsulatorElementR,InsulatorElementL,InsulatorElementM= Model(E,Fy1,Fy2,Imperfection,Density, Ei,Di,rhoi,Ti,Li, Ec,Dc,rhoc, Fyc, Lc)
        W2_strcture= Gravity(rhoc, Lc, cableR, cableM,cableL)
        ok= static_wind_Analysis(WindSpeed, cableR, cableM,cableL)
        # GetOutputs(1)
        # PlotTowerDisplacement(1)
        dispx.append(op.nodeDisp(310)[0])
        dispy.append(op.nodeDisp(284)[1])
        Rnode=[1995, 2085, 2176, 2266]
        Reaction0=[]
        Reaction90=[]
        op.reactions()
        for RR in Rnode:
            Reaction0.append(abs(op.nodeReaction(RR,1)))
            Reaction90.append(abs(op.nodeReaction(RR,2)))
        baseshear0=sum(Reaction0)/1000
        baseshear90=sum(Reaction90)/1000
        reactx.append(baseshear0)
        reacty.append(baseshear90)
        wind.append(WindSpeed*10)

        op.wipe()


    with open('combined_data.txt', 'w') as file:
        for x, y, rx, ry, w in zip(dispx, dispy, reactx, reacty, wind):
            file.write(f"{x} {y} {rx} {ry} {w}\n")

    dispx = np.array(dispx)
    dispy = np.array(dispy)
    reactionx = np.array(reactx)
    reactiony = np.array(reacty)
    resultant_disp = np.sqrt(dispx**2 + dispy**2)
    resultant_reaction = np.sqrt(reactionx**2 + reactiony**2)
    wind = np.array(wind)
    plt.figure(figsize=(8, 8))
    plt.plot(dispx, wind, 'o-', color='black', label='Wind vs. Disp X')
    plt.xlabel('displacement X (m)')
    plt.ylabel('Wind (m/s)', color='black')
    plt.tick_params('y', colors='black')
    plt.grid(True)
    ax1_right = plt.twinx()
    ax1_right.set_ylim([min(reactionx), max(reactionx)])
    ax1_right.set_ylabel('Reaction X (kN)', color='black')
    ax1_right.tick_params('y', colors='black')
    # plt.title('Wind vs. Disp X with Reaction X ')
    plt.show()

    # Plot 2: Wind vs. Disp Y with Reaction Y Range
    plt.figure(figsize=(8, 8))
    plt.plot(dispy, wind, 'o-', color='black', label='Wind vs. Disp Y')
    plt.xlabel('displacement Y (m)')
    plt.ylabel('Wind (m/s)', color='black')
    plt.tick_params('y', colors='black')
    plt.grid(True)
    ax2_right = plt.twinx()
    ax2_right.set_ylim([min(reactiony), max(reactiony)])
    ax2_right.set_ylabel('Reaction Y (kN)', color='black')
    ax2_right.tick_params('y', colors='black')
    # plt.title('Wind vs. displacement Y with Reaction Y Range')
    plt.show()

    # Plot 3: Wind vs. Resultant Disp with Resultant Reaction Range
    plt.figure(figsize=(8, 8))
    plt.plot(resultant_disp, wind, 'o-', color='black', label='Wind vs. Resultant Disp')
    plt.xlabel('Resultant displacement (m)')
    plt.ylabel('Wind (m/s)', color='black')
    plt.tick_params('y', colors='black')
    plt.grid(True)
    ax3_right = plt.twinx()
    ax3_right.set_ylim([min(resultant_reaction), max(resultant_reaction)])
    ax3_right.set_ylabel('Resultant Reaction (kN)', color='black')
    ax3_right.tick_params('y', colors='black')
    # plt.title('Wind vs. Resultant Disp with Resultant Reaction Range')
    plt.show()
    
    return

#Functions
def MainNodesTower(Lc, Tower_No):
    for i in range(len(Nodeiloc)):
        df = pd.read_excel('Input/NodeDomain.xlsx', sheet_name=Nodeiloc[i], header=0)
        for j in range(len(df['node'])):
            Tag=Tower_No*NumNodeDomain+ int(df['node'][j])
            op.node(Tag,float(df['x'][j]),float(df['y'][j])+Lc*Tower_No,float(df['z'][j]))
        
    return
def TowerReactionNode():
    Rnode =[]
    for i in op.getNodeTags():
        if op.nodeCoord(i)[2]==0:
            Rnode.append(i)
    return Rnode
global NintegrationS; global numSubdivJK1;global numSubdivJK1;global numSubdivIJ2; global numSubdivJK2
NintegrationS=1
numSubdivIJ1, numSubdivJK1 ,numSubdivIJ2,numSubdivJK2 = 3*NintegrationS, 2*NintegrationS, 2*NintegrationS,3*NintegrationS
def Section(E):
    
    with open('Input/SectionInfos.txt') as f4: 
        SectionsInfos=[]
        SectionsInfos=[list(map(float, line.split())) for line in f4]
    
    for SecTag, matTag, w, t, A in SectionsInfos:
        SecTag=int(SecTag)
        matTag=int(matTag)
        
        #Units are in Meter
        
        J = (w * t**3) / 3
        G = E / (2.0 * (1.0 + 0.3))
        Center=( w*t*t/2+ (w-t)*t*((w-t)/2+t) )/A #Center
        GJ=G*J
        I1x,I1y= -Center, -Center
        J1x,J1y= w-Center, -Center
        K1x,K1y=w-Center, -Center+t
        L1x, L1y= -Center,-Center+t
           
        I2x, I2y= -Center,   -Center+t
        J2x, J2y= -Center+t, -Center+t
        K2x, K2y=-Center+t,   w-Center
        L2x, L2y= -Center,    w-Center
        
        
        
        op.section( 'Fiber',	SecTag,'-GJ',GJ)
        op.patch('quad',	matTag,	numSubdivIJ1, numSubdivJK1,  I1x, I1y,J1x,J1y,K1x,K1y,L1x,L1y) 
        op.patch('quad',	matTag,	numSubdivIJ2, numSubdivJK2,  I2x, I2y,J2x,J2y,K2x,K2y,L2x,L2y) 
            
        if plotsection=='on':
            
            MySections = [['section', 'Fiber',	SecTag,'-GJ',GJ],
                          ['patch', 'quad',	matTag,	numSubdivIJ1, numSubdivJK1,  I1x, I1y,J1x,J1y,K1x,K1y,L1x,L1y],
                          ['patch','quad',	matTag,	numSubdivIJ2, numSubdivJK2,  I2x, I2y,J2x,J2y,K2x,K2y,L2x,L2y]]
            matcolor = ['r', 'lightgrey', 'gold', 'w', 'w', 'w']             
            vis.plot_fiber_section(MySections, matcolor=matcolor,fillflag=0)
    return

def Material(E, Fy1, Fy2, Ei,Ec, Fyc, Fyi):
    
        
    b1= 0.023
    op.uniaxialMaterial('Steel01', 1, Fy1, E, b1)
    # op.uniaxialMaterial('ElasticPP', 1, E, 10*E/Fy1)
    # op.uniaxialMaterial('ElasticPP', 1, E, Fy1/E)
    # op.uniaxialMaterial('Hardening', 1, E, Fy1, 0,b1*Fy1)
    b2= 0.02
    op.uniaxialMaterial('Steel01', 2, Fy2, E, b2)
    # op.uniaxialMaterial('ElasticPP', 2, E, 10*E/Fy2)
    # op.uniaxialMaterial('ElasticPP', 2, E, Fy2/E)
    # op.uniaxialMaterial('Hardening', 2, E, Fy2, 0, b2*Fy2)
    r=0.0254
    #Slippage1
    p1=[86.6e3,0.29/1000]
    p2=[197.3e3,2.19/1000]
    p3=[317e3,5.14/1000]
    n1=[-86.6e3,-0.29/1000]
    n2=[-197.3e3,-2.19/1000]
    n3=[-317e3,-5.14/1000]
    pinchX= 0.8;    pinchY= 0.2;    damage1= 0.0;    damage2= 0.0;    betaMUsteel= 0.0
    op.uniaxialMaterial('Hysteretic', 3, *p1, *p2, *p3, *n1, *n2, *n3, pinchX, pinchY, damage1, damage2, betaMUsteel)#1
    op.uniaxialMaterial('Elastic', 4, E*2) #2
    op.uniaxialMaterial('Elastic', 5, E*2) #3
    op.uniaxialMaterial('Elastic', 6, E*2) #4
    op.uniaxialMaterial('Elastic', 7, E*2) #6
    #moment zerolength
    
    pp1=[86.6e3*r,0.29/1000/r]
    pp2=[197.3e3*r,2.19/1000/r]
    pp3=[317e3*r,5.14/1000/r]
    nn1=[-86.6e3*r,-0.29/1000/r]
    nn2=[-197.3e3*r,-2.19/1000/r]
    nn3=[-317e3*r,-5.14/1000/r]
    op.uniaxialMaterial('Hysteretic', 8, *pp1, *pp2, *pp3, *nn1, *nn2, *nn3, pinchX, pinchY, damage1, damage2, betaMUsteel) #5


    #Insulator
    op.uniaxialMaterial('Steel01', 12, Fyi, Ei, Fyi/Ei)
    # op.uniaxialMaterial('Elastic', 12, Ec)
    # op.uniaxialMaterial('ElasticPP', 12, Ei, Fyi/Ei)
    op.uniaxialMaterial('InitStressMaterial',9,12, 0.45*Fyi)
    
    #Cable
    # op.uniaxialMaterial('Elastic', 10, Ec)
    # op.uniaxialMaterial('ElasticPP', 10, Ec, Fyc/Ec)
    op.uniaxialMaterial('Steel01', 10, Fyc, Ec, Fyc/Ec)
    op.uniaxialMaterial('InitStressMaterial',11,10, 0.45*Fyc)
    
    
    #Tower stiffness curve for supports at ends
    op.uniaxialMaterial('Hardening', 101,1674e3, 452e3, 0,-152e3)
    op.uniaxialMaterial('Hardening', 102,1095e3,230e3,0,-74e3)
    op.uniaxialMaterial('Hardening', 103,20150e3,3023e3,0,527e3)

    return
def Geometry(CoorI, CoorJ): 
    Coordinate_I= np.array(CoorI)
    Coordinate_J= np.array(CoorJ)
    
    L = np.sqrt(np.sum((Coordinate_J - Coordinate_I)**2))
    C= list((Coordinate_J-Coordinate_I)/L)
    vecxyNormal = [0, 1, 0.000001]
    vecxz =np.cross(C, vecxyNormal)
    return L , C, list(vecxz)
def Func_Geotransf (Type , nElem, vecxz):
    if Type == 'Beam':
        op.geomTransf('Linear', nElem, *vecxz)    
    if Type == 'Column':
        op.geomTransf('Corotational', nElem, *vecxz) 
    return
def floatingNodes():
    connectedNodes = []
    for ele in op.getEleTags():
        for nd in op.eleNodes(ele):
            connectedNodes.append(nd)
   
    definedNodes = op.getNodeTags()
    return list(set(connectedNodes) ^ set(definedNodes))

def generate_intermediate_coordinates(initial_coord, middle_coord, final_coord, num_segments, node, N,n):
    x_initial, y_initial, z_initial = initial_coord
    x_middle, y_middle, z_middle = middle_coord
    x_final, y_final, z_final = final_coord
    # Calculate the step size for each coordinate axis
    x_step = (x_final - x_initial) / num_segments
    y_step = (y_final - y_initial) / num_segments
    z_step_down = (z_middle - z_initial) / (num_segments // 2)
    z_step_up = (z_final - z_middle) / (num_segments // 2)
    # Generate intermediate coordinates with tags
    coordinates = []
    CelemEnds=[[node[0],x_initial, y_initial, z_initial]]
    CelemEnds=[]
    nodess=[]
    for i in range(1, num_segments):
        x_coordinate = x_initial + i * x_step
        y_coordinate = y_initial + i * y_step
        if i <= num_segments // 2:
            z_coordinate = z_initial + i * z_step_down
            node_tag=(n+1) *N + (n+1)*i
            nodess.append(node_tag)
            coordinates.append([node_tag, x_coordinate, y_coordinate, z_coordinate])
            CelemEnds.append([node_tag,x_coordinate,y_coordinate,z_coordinate])
        else:
            z_coordinate = z_middle + (i - num_segments // 2) * z_step_up
            node_tag=(n+1) *N + (n+1)* i
            nodess.append(node_tag)
            
            coordinates.append([node_tag, x_coordinate, y_coordinate, z_coordinate])
            CelemEnds.append([node_tag,x_coordinate,y_coordinate,z_coordinate])
        
    CelemEnds.append([node[1],x_final, y_final, z_final])
    return CelemEnds, nodess          
def CableElement(NtagCount, NelemTag, Insulators, rhoc, Dc, Ec, Lc, Sag, N_seg):  
    Ac=0.25*np.pi*Dc**2
    CableElementSaver=[]
    cableN=[]
    for n in range(len(Insulators)-1):
        node= [Insulators[n][0], Insulators[n+1][0]]
        initial_coord = Insulators[n][1:4]
        final_coord = Insulators[n + 1][1:4]
        middle_coord = [initial_coord[0], initial_coord[1]+ Lc/2 , initial_coord[2]- Sag]

        CelemEnds, nodess= generate_intermediate_coordinates(initial_coord, middle_coord, final_coord,
                                                                 N_seg, node, NtagCount,n)
        cableN.append(nodess)
        CableElementSaver.append(CelemEnds)
    CableElementSaver[0].insert(0, Insulators[0])
    for Tconductor in CableElementSaver:
        for iConductor in Tconductor:
            t, x,y,z= iConductor
            if t in op.getNodeTags():
                pass
            else:
                op.node(t, x,y,z)
    Connectivity=[]
    coorCable=[]
    L=len(CableElementSaver)
    for i in range(L):
        for t,x,y,z in CableElementSaver[i]:
            Connectivity.append(t)
            coorCable.append([x,y,z])
    mmd=[]
    for i in range(len(Connectivity)-1):
        mmd.append( [ Connectivity[i],Connectivity[i+1] ] )
    Ls=[]
    for i in range(len(Connectivity)-1):
        L_Cable , C, x=Geometry(coorCable[i], coorCable[i+1])
        Ls.append(L_Cable)
    elenode=[]
    Cable_Connectivity=[]
    for i,j in mmd:
        count=1
        elenode=[i,j]
        eletag=i+2*j+ 5*count+ N_seg*3
        Cable_Connectivity.append([eletag, *elenode])
        op.element('CorotTruss', eletag, *elenode, Ac, 11,'-rho', rhoc/9.81,'-doRayleigh', 0)

        count+=1    
    return Cable_Connectivity, cableN
def Insulator(N,Di, Li, rhoi, Insulators, matTag):
    NodesToC=[]
    InsulatorElement=[]
    n=0
    for node in Insulators:
        tag, x,y,z=node
        New_tag=int(tag)+N
        op.node(New_tag, x,y,z- Li)
        # op.fix(New_tag,1,0,0)
        eleTag= N+1000*(n+1)+Insulators[n][0]
        eleNodes=[tag, New_tag]
        NodesToC.append([New_tag, x,y,z- Li])
        A=0.25*np.pi*Di**2
        
        
        op.element('Truss', eleTag, *eleNodes, A, matTag,'-rho', rhoi/9.81, '-doRayleigh', 0)
        
        
        InsulatorElement.append(eleTag)
        n+=1
    return NodesToC, InsulatorElement

def GetOutputs(Step):
    if not os.path.exists(f"Outsteps_{Attempt}/{Step}"):
        os.mkdir(f"Outsteps_{Attempt}/{Step}")
    for Tower_No in range(iTower):
        workbook = openpyxl.load_workbook(f'Model_ColumnElements{Tower_No+1}.xlsx')
        for ii in range(len(C_Elementiloc)):
            sheet = workbook[C_Elementiloc[ii]]
            sheet[f'C{1}']="Stress Value (MPa)"
            sheet[f'D{1}']="Force Axial (N)"
            sheet[f'E{1}']="Mz(N-m)"
            sheet[f'F{1}']="My(N-m)"
            sheet[f'G{1}']="T(N-m)"

            for row in range(2, sheet.max_row + 1):
                # elementName = sheet.cell(row=row, column=1).value
                # SectionName = sheet.cell(row=row, column=2).value
                elementName =sheet[f'A{row}'].value
                SectionName=sheet[f'B{row}'].value
                y=-0.02549
                z=-0.02549
                stress=op.eleResponse(elementName,'section',str(3),'fiber', str(y), str(z), 'stress')[0]/1e6
                sheet[f'C{row}']=stress
                if op.eleResponse(elementName,'section', SectionName, 'force') !=[]:
                    sheet[f'D{row}']=op.eleResponse(elementName,'section', SectionName, 'force')[0]
                    sheet[f'E{row}']=op.eleResponse(elementName,'section', SectionName, 'force')[1]
                    sheet[f'F{row}']=op.eleResponse(elementName,'section', SectionName, 'force')[2]
                    sheet[f'G{row}']=op.eleResponse(elementName,'section', SectionName, 'force')[3]

        workbook.save(f'Outsteps_{Attempt}/{Step}/Model_ColumnElements{Tower_No+1}.xlsx')
        workbook.close()
            
        workbook = openpyxl.load_workbook(f'Model_BeamElements{Tower_No+1}.xlsx')
        for ii in range(len(B_Elementiloc)):
            sheet = workbook[B_Elementiloc[ii]]
            sheet[f'C{1}']="Stress Value (MPa)"
            sheet[f'D{1}']="Force Axial (N)"
            sheet[f'E{1}']="Mz(N-m)"
            sheet[f'F{1}']="My(N-m)"
            sheet[f'G{1}']="T(N-m)"
            
            for row in range(2, sheet.max_row + 1):
                elementName =sheet[f'A{row}'].value
                SectionName=sheet[f'B{row}'].value
                
                y=-0.02549
                z=-0.02549
                
                sheet[f'C{row}']=op.eleResponse(elementName,'section',str(3),'fiber', str(y), str(z), 'stress')[0]/1e6
                if op.eleResponse(elementName,'section', SectionName, 'force') !=[]:
                    sheet[f'D{row}']=op.eleResponse(elementName,'section', SectionName, 'force')[0]
                    sheet[f'E{row}']=op.eleResponse(elementName,'section', SectionName, 'force')[1]
                    sheet[f'F{row}']=op.eleResponse(elementName,'section', SectionName, 'force')[2]
                    sheet[f'G{row}']=op.eleResponse(elementName,'section', SectionName, 'force')[3]

            workbook.save(f'Outsteps_{Attempt}/{Step}/Model_BeamElements{Tower_No+1}.xlsx')
            workbook.close()
    return
def preprocess():
    if floatingNodes() != []: print("Following Nodes are disconnected:", floatingNodes())
    if testplot=='on':
        BSV.PlotModel(plotmode=3, draw_nodes=True, title='Initial Model Plot', 
                  onhover_message=True, fig_width=1000,fig_height=800,show_constrained=True,
                  plot_legends=True ,image_filename=f"Out_{Attempt}/Initial Model Plot")
    print("Model Preprocess is done!")

    return
#******************************************************************************************************************************
#Run Functions

def Pushover(WindSpeed, cableR, cableM,cableL):
    
    Tower_No=0
    ControlNode = 310+Tower_No*NumNodeDomain
    ControlNodeDof=1
    
    # PatTag=op.getPatterns()[-1]+1
    # op.timeSeries('Linear',PatTag)
    # op.pattern('Plain', PatTag,PatTag)
    # op.load(ControlNode, 0.0, 1., 0., 0., 0., 0.)
    
    # 8888888888888888888888888888888#
    staticWindloading(WindSpeed, cableR, cableM,cableL)
    # 8888888888888888888888888888888#
    
    du=0.001 ;    U=1 #m
    steps=int(U/du)
    
    op.constraints('Transformation')           
    op.numberer('RCM')  
    op.system('UmfPack')
    # op.system('Mumps','-ICNTL14', 40)
    # op.test('NormDispIncr', 1e-15, 150)
    op.test('FixedNumIter',200)
    op.algorithm('NewtonLineSearch',True, True, 0.8,1000, 0.1, 10.0)
    op.integrator("DisplacementControl", ControlNode, ControlNodeDof, du)
    op.analysis("Static")
    # 8888888888888888888888888888888#
    filename=f"Out_{Attempt}/Push_Wind_{WindSpeed*10}.out"
    op.recorder( 'Node', '-file', filename  , '-time','-node', 310,284, '-dof ', *DOFsToMatchc, 'disp')
    # 8888888888888888888888888888888#

   
    t1=time.time()
    for i in range(1, steps+1):
        # 8888888888888888888888888888888#
        ok= op.analyze(1)
        # 8888888888888888888888888888888#
        if i % 30 == 0:
            op.printModel('-file', f"OutPush_{Attempt}/PrintPushover{i}.txt")
            print(f"Step of plot:{i}")
            if Output=='on':
                GetOutputs(i)
                PlotTowerDisplacement(i)

        if ok<0:
            print(">>>>>>>>>>>step failed was:",i, f'Displacement was {i*du}m')
            break
    if ok<0:
        
        t2=time.time()
        print(f"********Pushover Analysis is Failed*********\n Took {(t2-t1)/60} Minutes to solve")
    else:
        t2=time.time()
        print(f"*********Pushover Analysis is Compeleted*********\n Took {(t2-t1)/60} Minutes to solve")
    
    return

def Modal(nMode):
    W2_strcture= op.eigen(nMode)
    op.modalProperties('-file', f"Out_{Attempt}/ModalAnalysis.txt")
    if modeShape=='on':
        for i in range(nMode):
            vfo.plot_modeshape(modenumber=i+1, scale=3000, filename=f'Out{Attempt}/mode{i}', overlap='yes')
    return W2_strcture
         
def reset_analysis():
    op.setTime(0.0)
    op.loadConst()
    op.wipeAnalysis() 
    return
def RunStaticWind(N):
    op.constraints('Transformation')
    # op.system('BandGen')
    op.system('UmfPack')
    # op.system('Mumps')
    # op.system('BandSPD')
    
    op.numberer('RCM')
    op.test('FixedNumIter',  200)
    
    op.integrator('LoadControl', 1/N)

    # op.algorithm ('RaphsonNewton')
    # op.algorithm('UmfPack')
    
    op.algorithm('NewtonLineSearch',True, True, 0.8,1000, 0.1, 10.0)
    # op.algorithm('Newton')
    op.analysis('Static','-noWarnings')
    result= op.analyze(N)
    
#     with ProcessPoolExecutor(max_workers=practiceCores) as exe:
#         for _ in range(N):
#             future = exe.submit(op.analyze, 1)
#             result = future.result()
    
    return result
def static_wind_Analysis(WindSpeedT, WindSpeedC, cableR,cableL):


    # t1=time.time()
    staticWindloading(WindSpeedT,WindSpeedC, cableR,cableL)
    ok=RunStaticWind(1)
    
    # t2=time.time()
    # if ok<0:
    #         print(f"Failed\n Took {round((t2-t1)/60,2)} Minutes to solve for Wind: {WindSpeed*10}")
    # else:
    #     print(f"*******Static Wind Load Analysis is done*******\n Took {round((t2-t1)/60,2)} Minutes to solve for Wind: {WindSpeed*10}")
    # if runplot=='on':
    #     BSV.PlotDefo(plotmode=3,scale_factor=1,onhover_message=True, plot_legends=True, vertical_axis=3, show_constrained=True,
    #                   fig_width=1200,fig_height=800,title='Deflected Shape After Static Wind Load', image_filename=f"Out_{Attempt}/Deflected Shape")
    # # op.printModel('-file', f"OutPush_{Attempt}/PrintPushover{'WindStatic'}.txt")
    # if Output=='on':
    # GetOutputs(1)
    # PlotTowerDisplacement(1)
    
    return ok
def RunStatic(N):
    # op.constraints('Plain')
    op.constraints('Transformation')
    op.system('UmfPack')
    # op.system('BandGen')
    # op.system('Mumps')
    # op.system('BandSPD')
    # op.numberer('Plain')
    op.numberer('RCM')
    op.test('FixedNumIter',  100)
    # op.test('RelativeNormDispIncr', 1.0e-15, 100)
    op.integrator('LoadControl', 1/N)

    # op.algorithm ('RaphsonNewton')
    # op.algorithm('UmfPack')
    # op.algorithm('NewtonLineSearch')
    op.algorithm('Newton')

    op.analysis('Static')
    
    ok= op.analyze(N)
    return ok
def Gravity(rhoc, Lc, cableR, cableL):
    t1=time.time()
    PatTag=1
    op.timeSeries( 'Linear', PatTag)
    op.pattern( 'Plain', PatTag, PatTag)
    for Tower_No in range(iTower):
        op.load( 196+Tower_No*NumNodeDomain, 0.0,0.0, -8944, 0.0, 0.0, 0.0)
        op.load( 195+Tower_No*NumNodeDomain, 0.0,0.0, -8944, 0.0, 0.0, 0.0)
    
    #Ground Wire Cable Load
    if cable=='on':
        PatTag=op.getPatterns()[-1]+1
        Wc=-rhoc*Lc/CableN_seg
        op.timeSeries( 'Linear', PatTag)
        op.pattern( 'Plain', PatTag, PatTag, '-fact', Wc)
        for Tower_No in range(iTower+1):
            for Cnode in cableR[Tower_No]:
                op.load(Cnode, 0,0, 1)
            for Cnode in cableL[Tower_No]:
                op.load(Cnode, 0,0, 1)

    # print("Gravity Loadings Appleid!")    
    # print("Gravity Analysis Started!")
    ok=RunStatic(1)
    if ok<0:
          print ("********Gravity Analysis Faild********")
    else:
        op.printModel('-file', f"OutPush_{Attempt}/PrintGravity.txt")
        t2=time.time()
    reset_analysis()
    print (f"********Gravity Analysis Completed*********\n Took {round((t2-t1)/60,2)} Minutes to solve")

    if runplot=='on':
        BSV.PlotDefo(plotmode=3,scale_factor=1,onhover_message=True, plot_legends=True, vertical_axis=3, show_constrained=True,
                      fig_width=800,fig_height=1200,title='Deflected Shape After Gravity', image_filename=f"Out_{Attempt}/Gravity Deflected Shape")
    if Runmodal=='on':
        W2_strcture= Modal(nMode)
        reset_analysis()
        return W2_strcture
    reset_analysis()
    return 0

def RunDynamic(w):
    # T1 = round(2*np.pi/w[0]**0.5,2)
    
    # Rayleigh damping
    # zi = 0.05 # Mode 1
    # zj = 0.03 # Mode 3
    # wi = w[0]**0.5
    # wj = w[2]**0.5
    # A = np.zeros((2,2))
    # A[0,0] = 1/wi; A[0,1] = wi
    # A[1,0] = 1/wj; A[1,1] = wj
    # b = np.zeros(2)
    # b[0] = zi
    # b[1] = zj
    # x = np.linalg.solve(0.5*A,b)
    # op.rayleigh(x[0],0,0,x[1])
    
    w = w[0]**0.5
    print(f'Structure Fundamental period, T = {2*np.pi/w} sec')
    dampRatio = 0.03
    op.rayleigh(0., 0., 0., 2*dampRatio*w)
    
       
    
    op.system('UmfPack')
    # op.system('Mumps')
    op.constraints('Transformation')
    op.numberer('RCM')
    # op.test('FixedNumIter',  150)
    # op.test('NormDispIncr', 0.0000001, 100)
    op.test('EnergyIncr', 1.0e-8, 100)
    op.integrator('Newmark', 0.5, 0.25)
    op.algorithm('KrylovNewton')
    op.analysis('Transient')
    print("Dynamic Analysis is set!")
    
    return
def ParallelRUN(duration):
    print(f"The duraion of time history Analysis is {duration} Seconds")
    nIncr= int(duration/timestep)
    t1 = time.time()
    success = True

    with ThreadPoolExecutor(max_workers=practiceCores) as exe:
        for _ in range(nIncr):
            future = exe.submit(op.analyze, 1, timestep)
            result = future.result()

            if result < 0:
                print("Analysis Failed!")
                success = False
                break
    t2 = time.time()

    if success:
        print(f">>> Analysis Completed! Took {round((t2-t1)/60,2)} Minutes to Run for {nIncr} steps")
    return
def dynamicWindLoading(WindSpeed,cableR,cableL,cableM):
    Factor = pd.read_excel('Input/LoadingData.xlsx', sheet_name='Factor', header=0)
    
    
    UcableT=11.65388961 #m/s
    UcableB=12.18667496 #m/s
    Nt1=9.65 #m/s
    Nt2=11.25 #m/s
    Nt3=12.26 #m/s
    
    #Tower
    with open('Input/TowerTimeseries.txt') as f: 
        u=[]
        u=[list(map(float, line.split())) for line in f]
    u=np.array(u)
    
    LoadilocX=['S1x','S2x','S3x']
    LoadilocY=['S1y','S2y','S3y']
    Utower=[Nt1,Nt2,Nt3]
    
    for i in range(len(Utower)):
        df1 = pd.read_excel('Input/LoadingData.xlsx', sheet_name=LoadilocX[i], header=0)
        df2 = pd.read_excel('Input/LoadingData.xlsx', sheet_name=LoadilocY[i], header=0)
        U2=(WindSpeed*Utower[i]+u[:,i])**2
        Fd=U2*Factor[AttackAngle][i]+U2*Factor[AttackAngle][i+3]
        
        #timeseries X
        PatTag=op.getPatterns()[-1]+1
        op.timeSeries('Path',PatTag,'-dt',timestep,'-values',*Fd, '-factor',np.cos(AttackAngle*np.pi/180)/len(df1['node']))
        op.pattern('Plain',PatTag,PatTag)
        for i in range(len(df1['node'])):
            op.load(int(df1['node'][i]), 1., 0.0, 0.0, 0.0, 0.0, 0.0)
        
        #timeseries Y
        PatTag=op.getPatterns()[-1]+1
        op.timeSeries('Path',PatTag,'-dt',timestep,'-values',*Fd, '-factor',np.sin(AttackAngle*np.pi/180)/len(df2['node']))
        op.pattern('Plain',PatTag,PatTag)
        for i in range(len(df2['node'])):
            op.load(int(df2['node'][i]), 0.0, 1., 0.0, 0.0, 0.0, 0.0)
        
    
    if cable=='on':
        with open('Input/CondTimeseries.txt') as f: 
            u=[]
            u=[list(map(float, line.split())) for line in f]
        ucond=np.array(u)
        arrays = np.split(ucond, [38, 76], axis=1)
        ucondR = arrays[0]
        ucondL = arrays[1]
        ucondM = arrays[2]

        for condi in range(iTower+1):

            for uti in range(int(len(ucondR[0])/2)):

                    V2=(WindSpeed*UcableB+ucondR[:,uti+condi])**2
                    Fd=V2*Factor[AttackAngle][6]
                    PatTag=op.getPatterns()[-1]+1
                    op.timeSeries('Path',PatTag,'-dt',timestep,'-values',*Fd)
                    op.pattern('Plain',PatTag,PatTag)
                    op.load(cableR[condi][uti], 1, 0., 0.)

            for uti in range(int(len(ucondL[0])/2)):

                    V2=(WindSpeed*UcableB+ucondL[:,uti+condi])**2
                    Fd=V2*Factor[AttackAngle][6]
                    PatTag=op.getPatterns()[-1]+1
                    op.timeSeries('Path',PatTag,'-dt',timestep,'-values',*Fd)
                    op.pattern('Plain',PatTag,PatTag)
                    op.load(cableL[condi][uti], 1, 0., 0.)

            for uti in range(int(len(ucondM[0])/2)):

                    V2=(WindSpeed*UcableT+ucondM[:,uti+condi])**2
                    Fd=V2*Factor[AttackAngle][6]
                    PatTag=op.getPatterns()[-1]+1
                    op.timeSeries('Path',PatTag,'-dt',timestep,'-values',*Fd)
                    op.pattern('Plain',PatTag,PatTag)
                    op.load(cableM[condi][uti], 1, 0., 0.)

    print("******************Force Timeseries applied******************")
    op.printModel('-file', "Currentmodel  with Dynamic Wind.txt")
    print("Model Printed!")
    return
def staticWindloading(WindSpeedT,WindSpeedC, cableR, cableL):
    Factor = pd.read_excel('Input/LoadingData.xlsx', sheet_name='Factor', header=0)
    Nt1=8.12 #m/s
    Nt2=9.63 #m/s
    Nt3=10.5 #m/s
    Utower=[Nt1,Nt2,Nt3]
    

    
    LoadilocX=['S1x','S2x','S3x']
    LoadilocY=['S1y','S2y','S3y']   
    Ft=[]
    for Tower_No in range(iTower):
        for j in range(3):
            df1 = pd.read_excel('Input/LoadingData.xlsx', sheet_name=LoadilocX[j], header=0)
            df2 = pd.read_excel('Input/LoadingData.xlsx', sheet_name=LoadilocY[j], header=0)
            
            V2=(WindSpeedT*Utower[j])**2
            Fd=(V2*Factor[AttackAngle][j]+V2*Factor[AttackAngle][j+3]) #*Solit[j]
            Ft.append(Fd)
            ForceX=Fd*np.cos(AttackAngle*np.pi/180)/len(df1['node'])
            ForceY=Fd*np.sin(AttackAngle*np.pi/180)/len(df2['node'])
            
            PatTag=op.getPatterns()[-1]+1
            op.timeSeries('Linear',PatTag)
            op.pattern("Plain", PatTag, PatTag,  '-fact', ForceX)
            
            for i in range(len(df1['node'])):
                op.load(int(df1['node'][i])+Tower_No*NumNodeDomain, 1, 0.0, 0.0, 0.0, 0.0, 0.0)
                
            PatTag=op.getPatterns()[-1]+1
            op.timeSeries('Linear',PatTag)
            op.pattern("Plain", PatTag, PatTag,  '-fact', ForceY*factorAngle) 
                
            for i in range(len(df2['node'])):
                op.load(int(df2['node'][i])+Tower_No*NumNodeDomain, 0.0, 1, 0.0, 0.0, 0.0, 0.0)
    
    if plotwindspeedprofile=='on':
    
        Z_values=[7.8601,21.8802,38.9301]
        plt.plot( Ft, Z_values, marker='o')
        plt.xlabel('F (N)')
        plt.ylabel('Z (m)')
        plt.title(f'{round(sum(Ft)/1000,2)}kN WindForce!')
        plt.grid(True)
        plt.show()
        print('WindLoad is ploted')
    if cable=='on':
    
        UcableB=10.71 #m/s
 
        for Tower_No in range(iTower+1):
            
            V2=(WindSpeedC*UcableB)**2
            Fd=V2*Factor[AttackAngle][6]
            # print(f'Cable Bottom:{Fd} newton')
            PatTag=op.getPatterns()[-1]+1
            op.timeSeries('Linear',PatTag)
            op.pattern("Plain", PatTag, PatTag, '-fact', Fd)
            
            for Cnode in cableR[Tower_No]:
                op.load(Cnode, 1, 0., 0.)
            for Cnode in cableL[Tower_No]:
                op.load(Cnode, 1, 0., 0.)
                

    return
#******************************************************************************************************************************
#Opensees Model Commands
def Model(E,Fy1,Fy2,Imperfection,Density, Ei,Di,rhoi,Ti,Li, Ec,Dc,rhoc, Fyc, Lc):
    t1=time.time()
    op.model('basic', '-ndm', 3,'-ndf', 6)
    Material(E, Fy1, Fy2, Ei,Ec, Fyc, Ti)
    Section(E)
    GEtag=1
    
    if Reliability=='off':
        NodesinTower = openpyxl.Workbook()
    counter=0
    Insulators_R=[]
    Insulators_L=[]
    
    for Tower_No in range(iTower):

        MainNodesTower(Lc, Tower_No)
        op.fix(1+Tower_No*NumNodeDomain, 1,1,1,1,1,1) 
        op.fix(2+Tower_No*NumNodeDomain, 1,1,1,1,1,1) 
        op.fix(3+Tower_No*NumNodeDomain, 1,1,1,1,1,1) 
        op.fix(4+Tower_No*NumNodeDomain, 1,1,1,1,1,1)
        
        #88888888888888888888888888888888888888888888888888888888888888888888888888888888888888888 BRACE ELEMENTS
        if Reliability=='off':
            Model_BeamElements = openpyxl.Workbook()
            
        for ii in range(len(B_Elementiloc)):

            if Reliability=='off':
                sheet1 = Model_BeamElements.create_sheet(title=f'{B_Elementiloc[ii]}')
                sheet1[f'A{1}']='Element Name'
                sheet1[f'B{1}']='Section Tag'
                sheet1[f'C{1}']='EndI'
                sheet1[f'D{1}']='EndJ'
                Count1=2
            df = pd.read_excel('Input/ElementDomain.xlsx', sheet_name=B_Elementiloc[ii], header=0)
            for jj in range(len(df['EleNum'])):
                nElem=Tower_No*NumNodeDomain+ int(df['EleNum'][jj])
                i=Tower_No*NumNodeDomain+ int(df['Endi'][jj])
                j=Tower_No*NumNodeDomain+ int(df['Endj'][jj])
                SecTag=int(df['SecTag'][jj])
                for secInfo in SectionsInfos:
                    if SecTag==secInfo[0]:
                        AreaS=secInfo[4]
                L,C, vecxz= Geometry(op.nodeCoord(i),op.nodeCoord(j))
                transTag=nElem
                integTag=nElem
                Func_Geotransf('Beam', transTag, vecxz)
                op.beamIntegration('Lobatto', integTag, SecTag, NintegrationB)
                Lseg=L/N_Cmesh
                A=[i]; fac=[0.5,1,0.5]
                for ii in range(N_Cmesh-1):
                    MeshnodeTag=op.getNodeTags()[-1]+1
                    Coors=list(np.array(op.nodeCoord(i)) + ((ii+1)*Lseg)*np.array(C)+ np.array(vecxz)*fac[ii]*Imperfection*L)
                    op.node(MeshnodeTag, *Coors)
                    A.append(MeshnodeTag)
                A.append(j)


                for idx in range(len(A)-1):
                    elemEnd=[A[idx],A[idx+1]]
                    op.element('dispBeamColumn',GEtag,*elemEnd ,transTag,integTag,'-mass',Density*AreaS)

                    if Reliability=='off':
                        sheet1[f'A{Count1}']=GEtag
                        sheet1[f'B{Count1}']=SecTag
                        sheet1[f'C{Count1}']=A[idx]
                        sheet1[f'D{Count1}']=A[idx+1]
                        Count1+=1
                    GEtag+=1
            if Reliability=='off':
                Model_BeamElements.save(f'Model_BeamElements{Tower_No+1}.xlsx')        
        #88888888888888888888888888888888888888888888888888888888888888888888888888888888888888888 LEG ELEMENTS
        if Reliability=='off':
            Model_ColumnElements= openpyxl.Workbook()

        for ii in range(len(C_Elementiloc)):
            if Reliability=='off':
                sheet2 = Model_ColumnElements.create_sheet(title=f'Zero{C_Elementiloc[ii]}')
                sheet2[f'A{1}']='Element Name'
                sheet2[f'B{1}']='LegIn Bet.'
                sheet2[f'C{1}']='EndI'
                sheet2[f'D{1}']='EndJ'
                Count2=2
                sheet3 = Model_ColumnElements.create_sheet(title=f'{C_Elementiloc[ii]}')
                sheet3[f'A{1}']='Element Name'
                sheet3[f'B{1}']='Section Tag'
                sheet3[f'C{1}']='EndI'
                sheet3[f'D{1}']='EndJ'
                Count3=2
            df = pd.read_excel('Input/ElementDomain.xlsx', sheet_name=C_Elementiloc[ii], header=0)
            for jj in range(len(df['EleNum'])):
                nElem=Tower_No*NumNodeDomain+ int(df['EleNum'][jj])


                i_orginal=Tower_No*NumNodeDomain+ int(df['Endi'][jj])
                j_orginal=Tower_No*NumNodeDomain+ int(df['Endj'][jj])
                i=op.getNodeTags()[-1]+1
                j=op.getNodeTags()[-1]+2

                SecTag=int(df['SecTag'][jj])
                for secInfo in SectionsInfos:
                    if SecTag==secInfo[0]:
                        AreaS=secInfo[4]
                op.node(i,*op.nodeCoord(Tower_No*NumNodeDomain+ int(df['Endi'][jj])))
                op.equalDOF(*[i_orginal,i], *DOFsToMatch)
                op.node(j,*op.nodeCoord(Tower_No*NumNodeDomain+ int(df['Endj'][jj])))
                op.equalDOF(*[j_orginal,j], *DOFsToMatch)

                op.element('zeroLength', GEtag, *[i_orginal,i], '-mat', *[3,4,5,6,8,7], '-dir', *DOFsToMatch)
                if Reliability=='off':
                    sheet2[f'A{Count2}'] =GEtag;            sheet2[f'B{Count2}']=nElem
                    sheet2[f'C{Count2}']=i_orginal;            sheet2[f'D{Count2}']=i
                    Count2+=1;            
                GEtag+=1

                op.element('zeroLength', GEtag, *[j_orginal,j], '-mat', *[3,4,5,6,8,7], '-dir', *DOFsToMatch)
                if Reliability=='off':
                    sheet2[f'A{Count2}']=GEtag;            sheet2[f'B{Count2}']=nElem
                    sheet2[f'C{Count2}']=j_orginal;            sheet2[f'D{Count2}']=j
                    Count2+=1;            
                GEtag+=1

                L,C, vecxz= Geometry(op.nodeCoord(i_orginal),op.nodeCoord(j_orginal))
                transTag=nElem
                integTag=nElem
                Func_Geotransf('Column', transTag, vecxz)
                op.beamIntegration('Lobatto', integTag, SecTag, NintegrationC)
                Lseg=L/N_Cmesh
                A=[i]; fac=[0.5,1,0.25]
                for ii in range(N_Cmesh-1):
                    MeshnodeTag=op.getNodeTags()[-1]+1
                    Coors=list(np.array(op.nodeCoord(i_orginal)) + ((ii+1)*Lseg)*np.array(C)+ np.array(vecxz)*fac[ii]*Imperfection*L)
                    op.node(MeshnodeTag, *Coors)
                    A.append(MeshnodeTag)
                A.append(j)


                for idx in range(len(A)-1):
                    elemEnd=[A[idx],A[idx+1]]
                    op.element('dispBeamColumn',GEtag,*elemEnd ,transTag,integTag,'-mass',Density*AreaS)

                    if Reliability=='off':
                        sheet3[f'A{Count3}']=GEtag
                        sheet3[f'B{Count3}']=SecTag
                        sheet3[f'C{Count3}']=A[idx]
                        sheet3[f'D{Count3}']=A[idx+1]
                        Count3+=1
                    GEtag+=1
        #88888888888888888888888888888888888888888888888888888888888888888888888888888888888888888 Conductor ELEMENTS

        if Reliability=='off':
            Model_ColumnElements.save(f'Model_ColumnElements{Tower_No+1}.xlsx')            
        
        Insulators_R.append([141+Tower_No*NumNodeDomain,*op.nodeCoord(141+Tower_No*NumNodeDomain)])
        Insulators_L.append([168+Tower_No*NumNodeDomain,*op.nodeCoord(168+Tower_No*NumNodeDomain)])    
        

        if Reliability=='off':
            sheet4 = NodesinTower.create_sheet(title=f'Tower{Tower_No+1}')
            sheet4['A1'] = "Node Tag"
            sheet4['B1'] = "x m"
            sheet4['C1'] = "y m"
            sheet4['D1'] = "z m"

            for i, tag in enumerate(op.getNodeTags()[counter:]):
                x,y,z= op.nodeCoord(tag)
                sheet4.cell(row=i+2, column=1, value=tag)
                sheet4.cell(row=i+2, column=2, value=x)
                sheet4.cell(row=i+2, column=3, value=y)
                sheet4.cell(row=i+2, column=4, value=z)
            counter=i+1+counter
            NodesinTower.save('NodesinTower.xlsx')  


    if cable=="on":

        op.model('basic', '-ndm', 3,'-ndf', 3)  

        ConnectingNodesR=[]
        ConnectingNodesL=[]
        

        AttachNode=[]
        
        for i, (tag, x, y, z) in enumerate(Insulators_R):
            New_tag = tag + int(1e5)
            
            ConnectingNodesR.append([New_tag, x, y, z])
            op.node(New_tag, x, y, z)
            op.equalDOF(Insulators_R[i][0], New_tag, *DOFsToMatchc)
            AttachNode.append(New_tag)
        for i, (tag, x, y, z) in enumerate(Insulators_L):
            New_tag = tag + int(1e5)
            ConnectingNodesL.append([New_tag, x, y, z])
            op.node(New_tag, x, y, z)
            op.equalDOF(Insulators_L[i][0], New_tag, *DOFsToMatchc)
            AttachNode.append(New_tag)


        N = int(2e5)
        nodesR,InsulatorElementR = Insulator(N, Di, Li, rhoi, ConnectingNodesR, 9)
        nodesL,InsulatorElementL = Insulator(N, Di, Li, rhoi, ConnectingNodesL, 9)


        #Defining the boundary cable
        nodesR=[*nodesR, list(np.array(nodesR[-1])+np.array([10000,0,Lc,0]))]
        nodesR=[list(np.array(nodesR[0])+np.array([100000,0,-1*Lc,0])), *nodesR]
        nodesR=[[int(sublist[0])] + sublist[1:] for sublist in nodesR]
        
        nodesL=[*nodesL, list(np.array(nodesL[-1])+np.array([10000,0,Lc,0]))]
        nodesL=[list(np.array(nodesL[0])+np.array([100000,0,-1*Lc,0])), *nodesL]
        nodesL=[[int(sublist[0])] + sublist[1:] for sublist in nodesL]


        for i in range(len(nodesR)):
            if i==0:
                op.node(*nodesR[i])
                tagEnd,x,y,z=nodesR[i]
                newTag=op.getNodeTags()[-1]+10
                op.node(newTag,x,y,z)
                op.fix(newTag,1,1,1)

                ZeroElemEnd=[newTag,tagEnd]
                op.element('zeroLength', op.getEleTags()[-1]+1, *ZeroElemEnd, '-mat', *[101,102,103], '-dir', *DOFsToMatchc)

            if i==len(nodesR)-1:
                op.node(*nodesR[i])
                tagEnd,x,y,z=nodesR[i]
                newTag=op.getNodeTags()[-1]+11
                op.node(newTag,x,y,z)
                op.fix(newTag,1,1,1)

                ZeroElemEnd=[newTag,tagEnd]
                op.element('zeroLength', op.getEleTags()[-1]+1, *ZeroElemEnd, '-mat', *[101,102,103], '-dir', *DOFsToMatchc)

        for i in range(len(nodesL)):
            if i==0:
                op.node(*nodesL[i])
                tagEnd,x,y,z=nodesL[i]
                newTag=op.getNodeTags()[-1]+12
                op.node(newTag,x,y,z)
                op.fix(newTag,1,1,1)

                ZeroElemEnd=[newTag,tagEnd]
                op.element('zeroLength', op.getEleTags()[-1]+1, *ZeroElemEnd, '-mat', *[101,102,103], '-dir', *DOFsToMatchc)


            if i==len(nodesL)-1:
                op.node(*nodesL[i])
                tagEnd,x,y,z=nodesL[i]
                newTag=op.getNodeTags()[-1]+13
                op.node(newTag,x,y,z)
                op.fix(newTag,1,1,1)

                ZeroElemEnd=[newTag,tagEnd]
                op.element('zeroLength', op.getEleTags()[-1]+1, *ZeroElemEnd, '-mat', *[101,102,103], '-dir', *DOFsToMatchc)

       
        Nnode1=int(3e5)
        Nnode2=int(3.02e5)
        
        N1 = int(6.0e5)
        N2 = int(7.0e5)
        
        
        Cable_Connectivity_R, cableR= CableElement(Nnode1,N1, nodesR, rhoc,Dc,Ec,Lc, Sag, CableN_seg)
        Cable_Connectivity_L, cableL= CableElement(Nnode2,N2, nodesL, rhoc,Dc,Ec,Lc, Sag, CableN_seg)
        

        CelementR = [eleTag for eleTag, _, _ in Cable_Connectivity_R]
        CelementL = [eleTag for eleTag, _, _ in Cable_Connectivity_L]
        

        CableNodesR = [i for eletag, i, j in Cable_Connectivity_R]
        CableNodesR.append(Cable_Connectivity_R[len(Cable_Connectivity_R)-1][2])
        CableNodesL = [i for eletag, i, j in Cable_Connectivity_L]
        CableNodesL.append(Cable_Connectivity_L[len(Cable_Connectivity_L)-1][2])
        
    t2=time.time()
    
    if cable=='on':
        print(f'Took {round((t2-t1)/60,2)}minutes to build the model!')
        return CelementR, CelementL, cableR, cableL,InsulatorElementR,InsulatorElementL
    else:
        print(f'Took {round((t2-t1)/60,2)}minutes to build the model!')
        return
#***********************************************************************************************************************# Recorders Function

# Recorders Function
def Recoders_DispMainNodes(runtype,Tower_No):
    for i in range(len(Nodeiloc)):
        df = pd.read_excel('Input/NodeDomain.xlsx', sheet_name=Nodeiloc[i], header=0)
        filename1=f"Out_{Attempt}/Tower_{Tower_No+1}_{runtype}_Nodes_in_Part_{Nodeiloc[i]}.out"
        
        RecNode=[]
        for i in list(df['node']):
            RecNode.append(i+ Tower_No*NumNodeDomain)
        
        op.recorder('Node', '-file', filename1,'-time', '-dT', Record_deltaT,'-node', *RecNode,'-dof ',*DOFsToMatch, 'disp')
        
    return
def Recorders_ReactionNodes(runtype,Tower_No):
    Rnode=[1995+NumNodeDomain*Tower_No, 2085+NumNodeDomain*Tower_No, 2176+NumNodeDomain*Tower_No, 2266+NumNodeDomain*Tower_No]
    filename1=f"Out_{Attempt}/Tower_NodeReaction{Tower_No+1}_{runtype}.out"
    op.reactions()
    op.recorder( 'Node', '-file', filename1  , '-time','-dT', Record_deltaT,'-node', *Rnode ,'-dof ', *DOFsToMatch, 'reaction')
    
    return
def Recorders_Cable(runtype,CelementR, CelementL, CelementM, InsulatorElementR,InsulatorElementL,InsulatorElementM, cableR, cableM,cableL):
    op.recorder( 'Element', '-file', f"Out_{Attempt}/CableforcesR_{runtype}.out", '-time',  '-ele', *CelementR, 'force')
    op.recorder( 'Element', '-file', f"Out_{Attempt}/CableforcesL_{runtype}.out", '-time',  '-ele', *CelementL, 'force')
    op.recorder( 'Element', '-file', f"Out_{Attempt}/CableforcesM_{runtype}.out", '-time',  '-ele', *CelementM, 'force')
    op.recorder( 'Element', '-file', f"Out_{Attempt}/InsulatorElementR_{runtype}.out", '-time',  '-ele', *InsulatorElementR, 'force')
    op.recorder( 'Element', '-file', f"Out_{Attempt}/InsulatorElementL_{runtype}.out", '-time',  '-ele', *InsulatorElementL, 'force')
    op.recorder( 'Element', '-file', f"Out_{Attempt}/InsulatorElementM_{runtype}.out", '-time',  '-ele', *InsulatorElementM, 'force')
    
    for i in range(iTower+1):
        op.recorder('Node', '-file', f"Out_{Attempt}/DispLineR{i+1}_{runtype}.out", '-time','-node', *cableR[i], '-dof', *DOFsToMatchc, 'disp')
    for i in range(iTower+1):
        op.recorder('Node', '-file', f"Out_{Attempt}/DispLineL{i+1}_{runtype}.out", '-time','-node', *cableL[i], '-dof', *DOFsToMatchc, 'disp')        
    for i in range(iTower+1):
        op.recorder('Node', '-file', f"Out_{Attempt}/DispLineM{i+1}_{runtype}.out", '-time','-node', *cableM[i], '-dof', *DOFsToMatchc, 'disp')
    return

#***********************************************************************************************************************
def recorder(runtype,CelementR, CelementL, CelementM, cableR, cableM,cableL,InsulatorElementR,InsulatorElementL,InsulatorElementM):
    op.initialize()
    for Tower_No in range(iTower):
        Recoders_DispMainNodes(runtype,Tower_No)
        Recorders_ReactionNodes(runtype,Tower_No)
    if cable=="on":
        Recorders_Cable(runtype,CelementR, CelementL, CelementM, InsulatorElementR,InsulatorElementL,InsulatorElementM, cableR, cableM,cableL)

    #Elements
    y = -0.03
    z = 0.02
    b=[]; c=[]
    for Tower_No in range(iTower):
        workbook = openpyxl.load_workbook(f'Model_ColumnElements{Tower_No+1}.xlsx')
        for ii in range(len(C_Elementiloc)):
            sheet=workbook[C_Elementiloc[ii]]
            for row in range(2, sheet.max_row + 1):
                b.append(sheet.cell(row=row, column=1).value)                                                             
            filename1=f"Out_{Attempt}/Tower_{Tower_No+1}_{C_Elementiloc[ii]}_Force.out"
            
            filename3=f"Out_{Attempt}/Tower_{Tower_No+1}_{C_Elementiloc[ii]}_Stress.out"
            op.recorder('Element', '-file', filename1,'-time', '-dT', Record_deltaT,'-ele', *b, 'force')
            
            op.recorder('Element', '-file', filename3,'-time', '-dT', Record_deltaT,'-ele', *b, 'section', str(3), 'fiber', str(y),str(z), 'stress')
        workbook.close()
        workbook = openpyxl.load_workbook(f'Model_BeamElements{Tower_No+1}.xlsx')
        for ii in range(len(B_Elementiloc)):
            sheet=workbook[B_Elementiloc[ii]]
            for row in range(2, sheet.max_row + 1):
                c.append(sheet.cell(row=row, column=1).value)                                                             
            filename1=f"Out_{Attempt}/Tower_{Tower_No+1}_{B_Elementiloc[ii]}_Force.out"
            
            filename3=f"Out_{Attempt}/Tower_{Tower_No+1}_{B_Elementiloc[ii]}_Stress.out"
            op.recorder('Element', '-file', filename1,'-time', '-dT', Record_deltaT,'-ele', *c, 'force')
            
            op.recorder('Element', '-file', filename3,'-time', '-dT', Record_deltaT,'-ele', *c, 'section', str(3), 'fiber', str(y),str(z), 'stress')
        workbook.close()
    return
#***********************************************************************************************************************
#Reliability
def NormalDist(Nsamples, mean, COV, plotDist='off'):
    sigma = COV * mean 
    x_values = np.linspace(norm.ppf(0.001, mean, sigma), norm.ppf(0.999, mean, sigma), Nsamples)
    random_variables = norm.rvs(mean, sigma, size=Nsamples)
      
    
    pdf_values = norm.pdf(x_values, mean, sigma)
    cdf_values = norm.cdf(x_values, mean, sigma)
    hazard_function_values = pdf_values / (1 - cdf_values)
    
    
    if plotDist=='on':
        fig, axs = plt.subplots(1, 3, figsize=(30, 10))
        # PDF
        axs[0].plot(x_values, pdf_values, 'r-', lw=2, label='PDF')
        axs[0].set_title('Probability Density Function')
        axs[0].set_xlabel('Value')
        axs[0].set_ylabel('Density')
        axs[0].legend()
        # CDF
        axs[1].plot(x_values, cdf_values, 'b-', lw=2, label='CDF')
        axs[1].set_title('Cumulative Distribution Function')
        axs[1].set_xlabel('Value')
        axs[1].set_ylabel('Probability')
        axs[1].legend()
        # Plot Hazard Function
        axs[2].plot(x_values, hazard_function_values, 'g-', lw=2, label='Hazard Function')
        axs[2].set_title('Hazard Function')
        axs[2].set_xlabel('Value')
        axs[2].set_ylabel('Hazard Rate')
        axs[2].legend()
        
        plt.tight_layout()
        plt.show()
        return random_variables
def LogNormalDist(Nsamples, mean, COV, plotDist='off'):
    sigma = math.sqrt(math.log(COV**2 + 1))
    mu = math.log(mean) - 0.5 * sigma**2
    x_values = np.linspace(lognorm.ppf(0.001, sigma, scale=np.exp(mu)), 
                           lognorm.ppf(0.999, sigma, scale=np.exp(mu)), Nsamples)
    
    random_variables = lognorm.rvs(sigma, scale=np.exp(mu), size=Nsamples)
    
    pdf_values = lognorm.pdf(x_values, sigma, scale=np.exp(mu))
    cdf_values = lognorm.cdf(x_values, sigma, scale=np.exp(mu))
    hazard_function_values = pdf_values / (1 - cdf_values)
    
    if plotDist == 'on':
        fig, axs = plt.subplots(1, 3, figsize=(30, 10))
        # PDF
        axs[0].plot(x_values, pdf_values, 'r-', lw=2, label='PDF')
        axs[0].set_title('Probability Density Function')
        axs[0].set_xlabel('Value')
        axs[0].set_ylabel('Density')
        axs[0].legend()
        # CDF
        axs[1].plot(x_values, cdf_values, 'b-', lw=2, label='CDF')
        axs[1].set_title('Cumulative Distribution Function')
        axs[1].set_xlabel('Value')
        axs[1].set_ylabel('Probability')
        axs[1].legend()
        # Hazard Function
        axs[2].plot(x_values, hazard_function_values, 'g-', lw=2, label='Hazard Function')
        axs[2].set_title('Hazard Function')
        axs[2].set_xlabel('Value')
        axs[2].set_ylabel('Hazard Rate')
        axs[2].legend()
        plt.tight_layout()
        plt.show()
    return random_variables
def UniformDist(Nsamples, mean, COV, plotDist='off'):
    std_dev = COV * mean
    width = 2 * std_dev * np.sqrt(3)
    a = mean - width / 2
    b = mean + width / 2
    
    x_values = np.linspace(a, b, Nsamples)
    random_variables = uniform.rvs(a, b-a, size=Nsamples)
    
    pdf_values = uniform.pdf(x_values, a, b-a)
    cdf_values = uniform.cdf(x_values, a, b-a)
    hazard_function_values = np.full_like(x_values, 1/(b-a))
    
    if plotDist == 'on':
        fig, axs = plt.subplots(1, 3, figsize=(30, 10))
        # PDF
        axs[0].plot(x_values, pdf_values, 'r-', lw=2, label='PDF')
        axs[0].set_title('Probability Density Function')
        axs[0].set_xlabel('Value')
        axs[0].set_ylabel('Density')
        axs[0].legend()
        # CDF
        axs[1].plot(x_values, cdf_values, 'b-', lw=2, label='CDF')
        axs[1].set_title('Cumulative Distribution Function')
        axs[1].set_xlabel('Value')
        axs[1].set_ylabel('Probability')
        axs[1].legend()
        # Hazard Function
        axs[2].plot(x_values, hazard_function_values, 'g-', lw=2, label='Hazard Function')
        axs[2].set_title('Constant Hazard Function')
        axs[2].set_xlabel('Value')
        axs[2].set_ylabel('Hazard Rate')
        axs[2].legend()
        plt.tight_layout()
        plt.show()
    return random_variables

def RunModel(X):
    
    op.wipe()
    E=X[0]
    Fy1=X[1]
    Fy2=X[2]
    Imperfection=X[3]
    Density=X[4]
    Ei=X[5]
    Di=X[6]
    rhoi=X[7]
    Ti=X[8]
    Li=3.5
    Ec=X[9]
    Dc=X[10]
    rhoc=X[11]
    Fyc=X[12]
    Lc=300.
    WindSpeedT=X[13]
    WindSpeedC=X[14]
    CelementR, CelementL, cableR,cableL,InsulatorElementR,InsulatorElementL= Model(E,Fy1,Fy2,Imperfection,Density, Ei,Di,rhoi,Ti,Li, Ec,Dc,rhoc, Fyc, Lc)
    Gravity(rhoc, Lc, cableR, cableL)
    # byc=Pushover(WindSpeed, cableR, cableM,cableL)
    static_wind_Analysis(WindSpeedT, WindSpeedC, cableR, cableL)
    
    # h=31.242

    nFailTowerBuk, nFailTowerYld, nFailTowerUlt ,nFailPullout, nFailSlide,nFailCond,nFailinsult= [False]*7
    # G Tower Failure
    dx,dy,dz=op.nodeDisp(168)[:3]
    Vdisp= np.sqrt(dx**2+dy**2)
    if Vdisp> 0.12:
        nFailTowerBuk=True
    if Vdisp> 0.32:
        nFailTowerYld=True
    if Vdisp> 0.53:
        nFailTowerUlt=True

    #G Conductor
    CondForce= op.basicForce(900094)[0]
    insulForce= op.basicForce(301141)[0]
    
    if CondForce> 50e3:
        nFailCond=True
    if insulForce>15000:
        nFailinsult=True
        
    # G Foundation
    Rnode=[1, 2, 3, 4]; AllRz=[];AllRXY=[]
    op.reactions()
    for R in Rnode:
        Rx,Ry,Rz=op.nodeReaction(R)[:3]
        ResultantXY=np.sqrt(Rx**2+Ry**2)
        AllRz.append(Rz)
        AllRXY.append(ResultantXY)
    RRxymax=max(AllRXY)
    Rzmax=max(AllRz)            
    if Rzmax>458166*0.46:
        nFailPullout=True
    if RRxymax>44927*0.46:
        nFailSlide=True
        
    op.wipe()
    return [Vdisp, nFailTowerBuk, nFailTowerYld, nFailTowerUlt, Rzmax, nFailPullout, RRxymax, nFailSlide, CondForce, insulForce, nFailCond,nFailinsult]

def MCS(runparallel, nTrials, E, Fy1, Fy2, Imperfection, Density, Ei, Di, rhoi, Ti, Ec, Dc, rhoc, Fyc, WindSpeed):
    op.wipeReliability()
    print(f'Using {runparallel}nTrials:{nTrials}\n\n\nmodel Parameters:\n',f'E, Fy1, Fy2: {E/1e6, Fy1/1e6, Fy2/1e6},\nImperfection: {Imperfection}',
          f'\nSteel Density: {Density},\ninsulator:\nEi, Di, rhoi, Ti :{Ei/1e6, Di, rhoi, Ti/1e3}\nCable: {Ec/1e6, Dc, rhoc, Fyc/1e6}\nWindSpeed (m/s): {WindSpeed*10}')
    if os.path.exists('MCS_result.txt'):
        os.remove('MCS_result.txt')

    # Structural Random Variables    
    op.randomVariable(1, 'lognormal', '-mean', E, '-stdv', 0.06*E) #E
    op.randomVariable(2, 'lognormal', '-mean', Fy1, '-stdv', 0.1*Fy1) #Fy1
    op.randomVariable(3, 'lognormal', '-mean', Fy2, '-stdv', 0.1*Fy2) #Fy2
    op.randomVariable(4, 'uniform', '-mean', Imperfection, '-stdv', 0.192*Imperfection) #Imperfection
    op.randomVariable(5, 'normal', '-mean', Density, '-stdv', 0.075*Density) #Density

    op.randomVariable(6, 'lognormal', '-mean', Ei, '-stdv', 0.06*Ei) #Ei
    op.randomVariable(7, 'uniform', '-mean', Di, '-stdv', 0.095*Di) #Di
    op.randomVariable(8, 'normal', '-mean', rhoi, '-stdv', 0.075*rhoi) #Density Insulator
    op.randomVariable(9, 'lognormal', '-mean', Ti, '-stdv', 0.1*Ti) #Ti

    op.randomVariable(10, 'lognormal', '-mean', Ec, '-stdv', 0.06*Ec) #Ec
    op.randomVariable(11, 'uniform', '-mean', Dc, '-stdv', 0.113*Dc) #Dc
    op.randomVariable(12, 'normal', '-mean', rhoc, '-stdv', 0.115*rhoc) #Density Cable
    op.randomVariable(13, 'lognormal', '-mean', Fyc, '-stdv', 0.1*Fyc) #Tc
    #***********************************************************************************************************************
    # Wind RVs
    # MRI wind speed Probability of Exceedance
    # WindPDF=-310.4256*np.e(-0.282*wind)

    op.randomVariable(14, 'lognormal', '-mean', WindSpeed, '-stdv', 0.053*WindSpeed) #Tower
    op.randomVariable(15, 'lognormal', '-mean', WindSpeed, '-stdv', 0.053*WindSpeed) #Cable
    
    t1ALL= time.time()
    nrv= len(op.getRVTags())
    op.probabilityTransformation('Nataf')

    AllX=[]
    for i in range(nTrials):    
        U=list(norm.rvs(size=nrv))
        X=op.transformUtoX(*U)
        AllX.append(X)

    if runparallel=='Process':
        with ProcessPoolExecutor(max_workers=practiceCores) as executor:
            results = list(executor.map(RunModel, AllX))
        t2ALL=time.time()
        print(f'********************Took {(t2ALL-t1ALL)/60} minutes To run Monte Carlo********************')
        return results, AllX

    if runparallel=='Thread':
        results=[]
        with ThreadPoolExecutor(max_workers=practiceCores) as exe:
            for X in AllX:
                future = exe.submit(RunModel, X)
                A=future.result()
                results.append(A)
        t2ALL=time.time()
        print(f'********************Took {(t2ALL-t1ALL)/60} minutes To run Monte Carlo********************')
        
        return results , AllX

    
    
def FORM():
    
    op.randomNumberGenerator('CStdLib')
    op.probabilityTransformation('Nataf','-print',3)
    op.reliabilityConvergenceCheck('Standard', '-e1', 1.0e-2, '-e2', 1e-2, '-print',1)
    op.functionEvaluator('Python', "-file", "opensees.analyze(1)")
    op.gradientEvaluator('Implicit')
    op.searchDirection('iHLRF')
    op.meritFunctionCheck('AdkZhang', "-multi", 2.0, "-add", 10.0,
                           "-factor", 0.5)
    op.stepSizeRule('Fixed', "-stepSize", 1.0)
    op.startPoint('Mean')
    op.findDesignPoint('StepSearch', "-maxNumIter", 15,
                        "-printDesignPointX", "designPointX.out")
    
    op.runFORMAnalysis('py_truss_FORM.out')
    
    return
def post_MCS(final_results, AllX):

    headerRes=["Vectordisp","FailBuckling","FailYielding","FailUltimate","Rzmax","FailPullout",
               "SlideShear","FailSlide","CondForce","insulForce","FailCond","nFailinsultor"]
    
    with open(f'results_{round(WindSpeed*10,3)}.txt', 'w') as file:
        for sublist in final_results:
            line = ', '.join(str(value) for value in sublist)
            file.write(line + '\n')    
    Results=pd.read_csv(f'results_{round(WindSpeed*10,3)}.txt', sep=',' ,header=None)
    Results.columns= headerRes
    
    with open('results.txt', 'w') as file:
        for sublist in final_results:
            line = ', '.join(str(value) for value in sublist)
            file.write(line + '\n')    
    header=["E","Fy1","Fy2","Imperfection","Density","Ei","Di","rhoi","Ti","Ec","Dc","rhoc","Fyc","WindSpeedT","WindSpeedC"]
    with open('RV_Vectors.txt', 'w') as file1:
        for X in AllX:
            line = ', '.join(str(value) for value in X)
            file1.write(line + '\n')
    RV_Vectors=pd.read_csv(f'RV_Vectors_{round(WindSpeed*10,3)}.txt', sep=',' ,header=None)
    RV_Vectors.columns= header
    
    RV_Vectors['WindSpeedT']*= 10
    RV_Vectors['WindSpeedC']*= 10
    
    
    return RV_Vectors, Results
# Nsamples=4000
# RV_Imperfection=NormalDist(Nsamples, Imperfection, 0.663)
# RV_rhoi=NormalDist(Nsamples, rhoi, 0.075)
# RV_rhoc=NormalDist(Nsamples, rhoc, 0.115)

# RV_Es=LogNormalDist(Nsamples, E, 0.06)
# RV_Fy1=LogNormalDist(Nsamples, 345e6, 0.1)
# RV_Fy2=LogNormalDist(Nsamples, 250e6, 0.1)

# RV_Es2=LogNormalDist(Nsamples, 0.02* 2e11, 0.25)
# RV_Ei=LogNormalDist(Nsamples, 1e11, 0.06)
# RV_Ec=LogNormalDist(Nsamples, 6.5e10, 0.06)


# RV_Fyi=LogNormalDist(Nsamples, 15000, 0.075)
# RV_Fyc=LogNormalDist(Nsamples, 162000, 0.075)

# RV_Di=UniformDist(Nsamples, 0.35, 0.095)
# RV_Dc=UniformDist(Nsamples, 0.025, 0.133)

# RV_windspeed=LogNormalDist(Nsamples,WindSpeed,0.0531)


#Static Analysis

# recorder('Gravity',CelementR, CelementL, CelementM, cableR, cableM,cableL,InsulatorElementR,InsulatorElementL,InsulatorElementM)
# W2_strcture= Gravity(rhoc, Lc, 1,1,1)
# recorder('Static',CelementR, CelementL, CelementM, cableR, cableM,cableL,InsulatorElementR,InsulatorElementL,InsulatorElementM)
# ok= static_wind_Analysis(WindSpeed, cableR, cableM,cableL)    





# Dynamic Analysis
# dynamicWindLoading()
# recorder('Dynamic',CelementR, CelementL, CelementM, cableR, cableM,cableL,InsulatorElementR,InsulatorElementL,InsulatorElementM)
# RunDynamic(W2_strcture)
# ParallelRUN(1)
#***********************************************************************************************************************
# CelementR, CelementL, CelementM, cableR, cableM,cableL,InsulatorElementR,InsulatorElementL,InsulatorElementM= Model(E,Fy1,Fy2,Imperfection,Density, Ei,Di,rhoi,Ti,Li, Ec,Dc,rhoc, Fyc, Lc)



# Pushover(WindSpeed, cableR, cableM,cableL)
op.wipe()

global Lc
Imperfection=0.075/100
E=210e9; Fy1= 345e6; Fy2= 250e6; Density=7850;
Di,Ei,Li,rhoi= 0.254, 100e9, 3, 5.68
Ai=0.25*np.pi*Di**2
Ti=15000/(Ai) #N/m2
Dc,Ec,Lc,rhoc, Tc, Sag= 0.025, 6.5e10, 335., 26.2, 50000, 0.0
Fyc=Tc/(0.25*np.pi*Dc**2)




# CelementR, CelementL, cableR, cableL,InsulatorElementR,InsulatorElementL= Model(E,Fy1,Fy2,Imperfection,Density, Ei,Di,rhoi,Ti,Li, Ec,Dc,rhoc, Fyc, Lc)
# W2_strcture= Gravity(rhoc, Lc, cableR, cableL)


WindSpeed=30 


if Reliability=='on':
    nTrials=5000
    runparallel='Process'
    final_results, AllX =MCS(runparallel, nTrials, E, Fy1, Fy2, Imperfection, Density, Ei, Di, rhoi, Ti, Ec, Dc, rhoc, Fyc, WindSpeed)
    RV_Vectors, Results= post_MCS(final_results, AllX)

